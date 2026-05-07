from flask import Flask, jsonify, request, make_response
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import os
import csv
import io

# Crear aplicación Flask independiente para FAB LAB
app = Flask(__name__)

# Configuración de BD SQLite
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///fablab_visitas.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'fablab-secret-key-2024'

# Inicializar BD
db = SQLAlchemy(app)

# Configurar CORS
CORS(app, resources={
    r"/api/*": {
        "origins": "*",
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type"]
    }
})

# ============= MODELO DE DATOS =============
class Visita(db.Model):
    __tablename__ = 'visitas'
    
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(120), nullable=False)
    rut = db.Column(db.String(20), nullable=False)
    correo = db.Column(db.String(120), nullable=False)
    tipo_visita = db.Column(db.String(80), nullable=False)
    telefono = db.Column(db.String(50), nullable=True)
    proposito = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'nombre': self.nombre,
            'rut': self.rut,
            'correo': self.correo,
            'tipo_visita': self.tipo_visita,
            'telefono': self.telefono,
            'proposito': self.proposito,
            'created_at': self.created_at.isoformat()
        }

# ============= RUTAS API =============

@app.route('/', methods=['GET'])
def home():
    return jsonify({
        "message": "FAB LAB INACAP - Sistema de Registro de Visitas",
        "versión": "1.0",
        "endpoints": {
            "crear_visita": "POST /api/visitas",
            "listar_visitas": "GET /api/visitas",
            "exportar_excel": "GET /api/visitas/export?format=excel",
            "exportar_csv": "GET /api/visitas/export?format=csv"
        }
    })

@app.route('/api/visitas', methods=['POST'])
def crear_visita():
    """Crear nuevo registro de visita"""
    data = request.get_json() or {}
    
    nombre = data.get('nombre', '').strip()
    rut = data.get('rut', '').strip()
    correo = data.get('correo', '').strip()
    tipo_visita = data.get('tipoVisita', '').strip() or data.get('tipo_visita', '').strip()
    telefono = data.get('telefono', '').strip()
    proposito = data.get('proposito', '').strip()
    
    # Validar campos obligatorios
    if not nombre or not rut or not correo or not tipo_visita:
        return jsonify({'error': 'Faltan campos obligatorios'}), 400
    
    try:
        visita = Visita(
            nombre=nombre,
            rut=rut,
            correo=correo,
            tipo_visita=tipo_visita,
            telefono=telefono,
            proposito=proposito
        )
        
        db.session.add(visita)
        db.session.commit()
        
        return jsonify({
            'message': 'Registro guardado exitosamente',
            'visita': visita.to_dict()
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/visitas', methods=['GET'])
def listar_visitas():
    """Obtener todas las visitas"""
    try:
        visitas = Visita.query.order_by(Visita.created_at.desc()).all()
        return jsonify([visita.to_dict() for visita in visitas]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/visitas/export', methods=['GET'])
def exportar_visitas():
    """Exportar visitas en Excel o CSV"""
    formato = request.args.get('format', 'csv').lower()
    
    try:
        visitas = Visita.query.order_by(Visita.created_at.desc()).all()
        
        if formato == 'excel':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Visitas"
                
                # Encabezados
                headers = ['ID', 'Nombre', 'RUT', 'Correo', 'Teléfono', 'Tipo de Visita', 'Propósito', 'Fecha de Registro']
                ws.append(headers)
                
                # Formato encabezado
                header_fill = PatternFill(start_color="ED1C24", end_color="ED1C24", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Datos
                for visita in visitas:
                    ws.append([
                        visita.id,
                        visita.nombre,
                        visita.rut,
                        visita.correo,
                        visita.telefono or '',
                        visita.tipo_visita,
                        visita.proposito or '',
                        visita.created_at.strftime('%d/%m/%Y %H:%M')
                    ])
                
                # Ajustar columnas
                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 25
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 30
                ws.column_dimensions['H'].width = 18
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response.headers['Content-Disposition'] = 'attachment; filename=visitas_fablab.xlsx'
                
                return response
            except ImportError:
                formato = 'csv'  # Fallback a CSV
        
        if formato == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['ID', 'Nombre', 'RUT', 'Correo', 'Teléfono', 'Tipo de Visita', 'Propósito', 'Fecha de Registro'])
            
            for visita in visitas:
                writer.writerow([
                    visita.id,
                    visita.nombre,
                    visita.rut,
                    visita.correo,
                    visita.telefono or '',
                    visita.tipo_visita,
                    visita.proposito or '',
                    visita.created_at.strftime('%d/%m/%Y %H:%M')
                ])
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            response.headers['Content-Disposition'] = 'attachment; filename=visitas_fablab.csv'
            
            return response
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.errorhandler(404)
def not_found(error):
    return jsonify({"error": "Endpoint no encontrado"}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({"error": "Error interno del servidor"}), 500

# ============= INICIALIZACIÓN =============

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        print("✓ Base de datos FAB LAB creada")
    
    print("🚀 Servidor FAB LAB iniciando...")
    print("📍 Escuchando en: http://0.0.0.0:5001 (accesible en la red local)")
    print("📍 CORS habilitado para todos los orígenes")
    print("📊 Documentación: GET http://<IP_DEL_SERVIDOR>:5001/")
    
    app.run(host='0.0.0.0', port=5001, debug=True)
